import pandas as pd
import openpyxl
from openpyxl import load_workbook
from typing import Dict, Any, List, Union
import os
import logging
from datetime import datetime

# Setup logger for this module
logger = logging.getLogger(__name__)

def parse_excel(file_path: Union[str, os.PathLike]) -> Dict[str, Any]:
    """
    Parses an Excel file (.xlsx or .xls) and extracts text from cells, sheets, and metadata.
    
    Args:
        file_path: Path to the Excel file.
        
    Returns:
        A dictionary containing extracted data including:
        - text from all sheets
        - sheet names and their content
        - metadata (author, creation date, etc.)
        - full text for LLM processing
    """
    extracted_data = {
        "sheets": [],
        "metadata": {},
        "full_text": ""
    }
    
    try:
        # Load workbook for metadata
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        # Extract metadata
        extracted_data["metadata"] = {
            "title": wb.properties.title,
            "author": wb.properties.creator,
            "created": wb.properties.created,
            "modified": wb.properties.modified,
            "last_modified_by": wb.properties.lastModifiedBy,
            "num_sheets": len(wb.sheetnames),
            "file_type": "excel"
        }
        
        # Process each sheet
        all_text_parts = []
        for sheet_name in wb.sheetnames:
            try:
                # Read sheet with pandas for better handling of merged cells and formatting
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                
                # Convert DataFrame to text, handling NaN values
                sheet_text = []
                for row in df.itertuples(index=False):
                    row_text = []
                    for cell in row:
                        if pd.notna(cell):  # Skip NaN/None values
                            cell_text = str(cell).strip()
                            if cell_text:
                                row_text.append(cell_text)
                    if row_text:  # Only add non-empty rows
                        sheet_text.append(" | ".join(row_text))
                
                sheet_content = "\n".join(sheet_text)
                
                # Store sheet data
                sheet_data = {
                    "sheet_name": sheet_name,
                    "text": sheet_content,
                    "row_count": len(df),
                    "column_count": len(df.columns)
                }
                extracted_data["sheets"].append(sheet_data)
                
                # Add to full text with sheet name as header
                if sheet_content.strip():
                    all_text_parts.append(f"=== Sheet: {sheet_name} ===\n{sheet_content}\n")
                
            except Exception as e:
                logger.warning(f"Error processing sheet '{sheet_name}': {str(e)}")
                continue
        
        # Combine all text
        extracted_data["full_text"] = "\n\n".join(all_text_parts)
        
        # Log extraction quality
        if not extracted_data["full_text"].strip():
            logger.warning("⚠️ No text could be extracted from the Excel file")
        else:
            logger.info(f"✅ Successfully extracted {len(extracted_data['full_text'])} characters from Excel file")
            logger.info(f"Processed {len(extracted_data['sheets'])} sheets")
        
        return extracted_data
        
    except Exception as e:
        logger.error(f"Error parsing Excel file {file_path}: {str(e)}")
        return {"error": str(e)}

if __name__ == '__main__':
    # Configure basic logging for direct script execution
    if not logging.getLogger().handlers:
        logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
    
    # Test the parser with a sample Excel file
    test_file = "example.xlsx"
    if os.path.exists(test_file):
        result = parse_excel(test_file)
        logger.info("\nParsed Excel Data:")
        logger.info(f"Number of sheets: {len(result.get('sheets', []))}")
        logger.info(f"Total text length: {len(result.get('full_text', ''))}")
        logger.info("\nMetadata:")
        for key, value in result.get('metadata', {}).items():
            logger.info(f"  {key}: {value}")
    else:
        logger.warning(f"Test file {test_file} not found. Create a sample Excel file to test the parser.") 